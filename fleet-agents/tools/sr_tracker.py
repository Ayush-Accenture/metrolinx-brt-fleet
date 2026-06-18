"""
SR Result Tracker — records SR creation results to fleet.xlsx and result.json,
then uploads all output artefacts to blob storage (fmi/out/).

fleet.xlsx format
-----------------
  Sheet name : BRT
  Columns    : Date | Task No. | Details

  Date       : SR raised date  (M/D/YYYY, e.g. "6/15/2026")
  Task No.   : SCTASK number   (e.g. "SCTASK0350947"); falls back to SR/RITM number
  Details    : Multi-line block:
                 Requestor Name - <name>
                 Action Requested - Monitoring Tool – Add/Remove from Device/Site Monitoring Alert Notification
                 Details of Request - TA: BRT
                 Note: This is device movement not deletion. Please do not remove these devices from SNOW.
                 List of BRT vehicles renamed/Moved from LTM to PROD:
                 <bus numbers, one per line, or NA>
                 List of BRT vehicles renamed/Moved from PROD to LTM:
                 <bus numbers, one per line, or NA>

Blob uploads (container: fmi, prefix: out/)
-------------------------------------------
  out/BRT_FleetMovement_{run_id}.xlsx
  out/fleet.xlsx
  out/result_{run_id}.json
"""

from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

log = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────
BLOB_OUT_PREFIX = "out"
FLEET_XLSX_NAME = "fleet.xlsx"
FLEET_XLSX_BLOB_PATH = f"{BLOB_OUT_PREFIX}/{FLEET_XLSX_NAME}"
SHEET_NAME = "BRT"

_FLEET_COLUMNS = ["Date", "Task No.", "Details"]
_HEADER_COLOR = "1A5276"   # dark blue


# ── Details block builder ────────────────────────────────────────────────────

def _build_details(requestor_name: str, recon_result: Any) -> str:
    """Build the multi-line Details cell content from reconciliation results."""
    def _is_ltm(folder: str) -> bool:
        return "ltm" in folder.lower()

    ltm_to_prod: list[str] = []
    prod_to_ltm: list[str] = []
    for item in recon_result.moved:
        bus = str(item.get("bus_number") or item.get("device", "?"))
        to_folder = item.get("to_folder", "")
        if _is_ltm(to_folder):
            prod_to_ltm.append(bus)
        else:
            ltm_to_prod.append(bus)

    ltm_to_prod_str = "\n".join(ltm_to_prod) if ltm_to_prod else "NA"
    prod_to_ltm_str = "\n".join(prod_to_ltm) if prod_to_ltm else "NA"

    return (
        f"Requestor Name - {requestor_name}\n"
        "Action Requested - Monitoring Tool – Add/Remove from Device/Site Monitoring Alert Notification\n"
        "Details of Request - TA: BRT\n"
        "Note: This is device movement not deletion. Please do not remove these devices from SNOW.\n"
        "List of BRT vehicles renamed/Moved from LTM to PROD:\n"
        f"{ltm_to_prod_str}\n"
        "List of BRT vehicles renamed/Moved from PROD to LTM:\n"
        f"{prod_to_ltm_str}"
    )


# ── Blob helpers ─────────────────────────────────────────────────────────────

def _get_container_client():
    """Return an Azure ContainerClient using available credentials (SAS → key → CLI)."""
    from azure.storage.blob import BlobServiceClient, ContainerClient

    if config.AZURE_STORAGE_SAS_URL:
        return ContainerClient.from_container_url(config.AZURE_STORAGE_SAS_URL)

    account_url = f"https://{config.AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
    if config.AZURE_STORAGE_ACCOUNT_KEY:
        svc = BlobServiceClient(account_url=account_url, credential=config.AZURE_STORAGE_ACCOUNT_KEY)
    else:
        from azure.identity import AzureCliCredential
        svc = BlobServiceClient(account_url=account_url, credential=AzureCliCredential())
    return svc.get_container_client(config.AZURE_STORAGE_CONTAINER_NAME)


def _upload_bytes(container_client, blob_path: str, data: bytes) -> str:
    """Upload bytes to blob (overwrite). Returns canonical blob URL (no SAS)."""
    container_client.get_blob_client(blob_path).upload_blob(data, overwrite=True)
    return (
        f"https://{config.AZURE_STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
        f"/{config.AZURE_STORAGE_CONTAINER_NAME}/{blob_path}"
    )


def _download_bytes(container_client, blob_path: str) -> bytes | None:
    """Download blob bytes; returns None if blob does not exist."""
    try:
        return container_client.get_blob_client(blob_path).download_blob().readall()
    except Exception:
        return None


# ── fleet.xlsx helpers ───────────────────────────────────────────────────────

def _load_or_create_fleet_workbook(raw: bytes | None) -> Workbook:
    """Load fleet.xlsx from bytes, or create a fresh one with styled header on 'BRT' sheet."""
    if raw:
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        # Ensure the BRT sheet exists (tolerate files created by earlier versions)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(_FLEET_COLUMNS)
            _style_header_row(ws)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(_FLEET_COLUMNS)
    _style_header_row(ws)
    return wb


def _style_header_row(ws) -> None:
    fill = PatternFill("solid", fgColor=_HEADER_COLOR)
    bold_white = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit_columns(ws) -> None:
    """Best-effort column width — capped at 80 for the Details column."""
    caps = {"Date": 14, "Task No.": 18, "Details": 80}
    for col in ws.columns:
        header = str(col[0].value or "")
        cap = caps.get(header, 60)
        max_len = max(
            (max(len(line) for line in str(c.value or "").splitlines()) if c.value else 0)
            for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, cap)


# ── Public API ───────────────────────────────────────────────────────────────

def record_sr_result(
    run_id: str,
    sr_number: str,
    ritm_number: str,
    sr_sys_id: str,
    sctask_number: str,
    recon_result: Any,
    output_workbook_path: str,
    requestor_name: str = "",
    run_summary: str = "",
) -> dict[str, Any]:
    """
    Record SR creation results to fleet.xlsx and result.json, then upload to blob.

    Parameters
    ----------
    run_id               : Pipeline run identifier (e.g. "BRT-2026-06-18")
    sr_number            : ServiceNow REQ number (e.g. "REQ0123456")
    ritm_number          : ServiceNow RITM number
    sr_sys_id            : ServiceNow internal sys_id
    sctask_number        : SCTASK number written to fleet.xlsx Task No. column
    recon_result         : core.reconciler.ReconcileResult (.moved / .unmoved / .unidentified)
    output_workbook_path : Local path to BRT_FleetMovement_*.xlsx
    requestor_name       : Display name of the requestor (shown in Details cell)
    run_summary          : Optional LLM narrative summary

    Returns
    -------
    dict with local paths and (when blob is live) blob URLs
    """
    now_utc = datetime.now(timezone.utc)
    agency = "BRT"
    wb_filename = os.path.basename(output_workbook_path)
    sr_status = "error" if sr_number.startswith("SR-ERROR-") else "success"

    account_name = config.AZURE_STORAGE_ACCOUNT_NAME
    container_name = config.AZURE_STORAGE_CONTAINER_NAME
    wb_blob_path = f"{BLOB_OUT_PREFIX}/{wb_filename}"
    wb_blob_url = (
        f"https://{account_name}.blob.core.windows.net/{container_name}/{wb_blob_path}"
    )
    fleet_blob_url = (
        f"https://{account_name}.blob.core.windows.net/{container_name}/{FLEET_XLSX_BLOB_PATH}"
    )

    os.makedirs(config.OUTPUT_DIR, exist_ok=True)

    # ── Write result.json ────────────────────────────────────────────────────
    result_data: dict[str, Any] = {
        "run_id": run_id,
        "agency": agency,
        "sr_number": sr_number,
        "ritm_number": ritm_number,
        "sr_sys_id": sr_sys_id,
        "sctask_number": sctask_number,
        "requestor_name": requestor_name,
        "sr_status": sr_status,
        "created_at_utc": now_utc.isoformat(),
        "fleet_movement_file": wb_filename,
        "fleet_movement_blob_path": wb_blob_path,
        "fleet_movement_blob_url": wb_blob_url,
        "fleet_tracking_blob_path": FLEET_XLSX_BLOB_PATH,
        "fleet_tracking_blob_url": fleet_blob_url,
        "moved_count": len(recon_result.moved),
        "unmoved_count": len(recon_result.unmoved),
        "unidentified_count": len(recon_result.unidentified),
        "moved": recon_result.moved,
        "unmoved": recon_result.unmoved,
        "unidentified": recon_result.unidentified,
        "run_summary": run_summary,
    }
    result_local_path = os.path.join(config.OUTPUT_DIR, f"result_{run_id}.json")
    with open(result_local_path, "w", encoding="utf-8") as fh:
        json.dump(result_data, fh, indent=2)
    log.info("result.json written: %s", result_local_path)

    # ── Update fleet.xlsx ────────────────────────────────────────────────────
    fleet_local_path = os.path.join(config.OUTPUT_DIR, FLEET_XLSX_NAME)

    if config.USE_MOCK_BLOB:
        raw = None
        if os.path.exists(fleet_local_path):
            with open(fleet_local_path, "rb") as fh:
                raw = fh.read()
        container_client = None
    else:
        container_client = _get_container_client()
        raw = _download_bytes(container_client, FLEET_XLSX_BLOB_PATH)
        if raw is None and os.path.exists(fleet_local_path):
            with open(fleet_local_path, "rb") as fh:
                raw = fh.read()

    fleet_wb = _load_or_create_fleet_workbook(raw)
    ws = fleet_wb[SHEET_NAME]

    # Date cell: M/D/YYYY  (no leading zeros, matching the example format)
    date_str = now_utc.strftime("%-m/%-d/%Y") if os.name != "nt" else now_utc.strftime("%#m/%#d/%Y")

    # Task No.: prefer SCTASK, fall back to RITM then REQ
    task_no = sctask_number or ritm_number or sr_number

    details_text = _build_details(requestor_name or "—", recon_result)

    # Append the new tracking row
    ws.append([date_str, task_no, details_text])
    last_row = ws.max_row

    # Wrap text and top-align the Details cell; auto-height via row_dimensions
    details_cell = ws.cell(row=last_row, column=3)
    details_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[last_row].height = max(15 * len(details_text.splitlines()), 60)

    _autofit_columns(ws)
    fleet_wb.save(fleet_local_path)
    log.info("fleet.xlsx updated (sheet=%s, row=%d): %s", SHEET_NAME, last_row, fleet_local_path)

    # ── Skip blob upload in mock mode ─────────────────────────────────────────
    if config.USE_MOCK_BLOB:
        log.info(
            "USE_MOCK_BLOB=true — blob upload skipped; files saved locally under %s",
            config.OUTPUT_DIR,
        )
        return {
            "result_json_local": result_local_path,
            "fleet_xlsx_local": fleet_local_path,
            "workbook_local": output_workbook_path,
            "blob_upload_skipped": True,
        }

    # ── Upload to blob (fmi/out/) ────────────────────────────────────────────
    result_blob_path = f"{BLOB_OUT_PREFIX}/result_{run_id}.json"
    urls: dict[str, str] = {}

    if os.path.exists(output_workbook_path):
        with open(output_workbook_path, "rb") as fh:
            urls["workbook_blob_url"] = _upload_bytes(container_client, wb_blob_path, fh.read())
        log.info("Uploaded %s → blob://%s/%s", wb_filename, container_name, wb_blob_path)
    else:
        log.warning("Workbook not found at %s — skipping blob upload", output_workbook_path)

    with open(fleet_local_path, "rb") as fh:
        urls["fleet_xlsx_blob_url"] = _upload_bytes(container_client, FLEET_XLSX_BLOB_PATH, fh.read())
    log.info("Uploaded fleet.xlsx → blob://%s/%s", container_name, FLEET_XLSX_BLOB_PATH)

    with open(result_local_path, "rb") as fh:
        urls["result_json_blob_url"] = _upload_bytes(container_client, result_blob_path, fh.read())
    log.info("Uploaded result_%s.json → blob://%s/%s", run_id, container_name, result_blob_path)

    return {
        "result_json_local": result_local_path,
        "fleet_xlsx_local": fleet_local_path,
        "workbook_local": output_workbook_path,
        **urls,
    }
